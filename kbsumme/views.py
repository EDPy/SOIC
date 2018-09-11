from django.shortcuts import render, redirect, get_object_or_404
from django.core.files.storage import FileSystemStorage
from django.utils import timezone
import pandas as pd
import os
import openpyxl


from .models import Posd, T3000db, KbMeta, Stueckliste, Pbb, PbbMeta, Customer
from .project_graph import createDetailProjectGraphic
from .io_statistic import iograph, total_IO
from dateutil.parser import parse

#TODO: Upload with choice of upload sheet: KBSUMME, Calculation, Stueckliste
#TODO: Before upload check all necesary fields, if okay.. include in database
#TODO: Delete database entry possibility

#global variables objects
t3000_inst = T3000db() # Same as post_inst
kbmeta_objects = KbMeta.objects
t3000_objects = T3000db.objects
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def home(request):
    #createDetailProjectGraphic(206019)

    '''
    query_total = t3000_objects.filter(posd__hgpos__regex=r'^(10000)\y')
    query_dict = {'totalcost': [0.0], 'projname':['Test'], 'customer':['SK'],
    'datecalc':['2017-01-01'], 'ExcelCalc': ['file'], 'dateupload':['2017-01-02'] }
    '''

    return render(request, 'kbsumme/home.html', {'kbmetaobj': kbmeta_objects.all(),})

def upPosd(request):
    ''' After form sumitted in html 'POST', it will be catched here and evaluated
    This method checks if file has been passed if yes it forwards it to upPosdFunc.
    Afterwards html is opened with updated information.

    If file was not selected, error message will occur.
    If browser is only refreshed (updated), content will be shown, no error.. See last return render
    '''
    posd_objects = Posd.objects  # Is an object to pass all objects to html to list items in table format

    if request.method == 'POST':
        if request.FILES:
            myfile = request.FILES['myfile']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name.replace(" ", ""), myfile)
            uploaded_file_url = fs.url(filename)
            r = {}
            r = upPosdFunc(request, uploaded_file_url, posd_objects)
            if r['code'] == 1:
                return render(request, 'kbsumme/upPosd.html', {
                'error':r['message'], 'posd_objects':posd_objects,
                })
            else:
                return render(request, 'kbsumme/upPosd.html', {
                'uploaded_file_url': uploaded_file_url,
                'success':'File successful loaded', 'posd_objects':posd_objects,})
        else:
            return render(request, 'kbsumme/upPosd.html', {
                'error':'Please select a file to upload', 'posd_objects': posd_objects,})
    else:
        return render(request, 'kbsumme/upPosd.html', {'posd_objects': posd_objects, })

def upPosdFunc(request, file, posd_objects):
    '''
    workbook will be opened, worsheet assigned. If not successfull error message will occur in the terminal.
    Otherwise all files are read into the Posd database. Filled in with data from the worksheet on pre-defined
    fields. Here: A2-A204, B2-B204, C2-C204, D2-D204, E2-E204
    '''

    r = {'code':0, 'message':''}
    try:
        wb = openpyxl.load_workbook(BASE_DIR + file, read_only=True)
        ws = wb['T1_general']
    except:
        r['code']=1
        r['message']='Could not load workbook or worksheet "T1_general"'
        return r

    posd_inst = Posd()  # Is an instance of class Posd to read into the database

    # Fetching excel sheet and stores into database.
    i=1 #Starts with number 2 to fetch data. Row 1 is the table header.
    try:
        while i < ws.max_row:
            i+=1
            hg = str(ws['A{}'.format(i)].value)

            if len(str(ws['B{}'.format(i)].value)) == 1:
                pos = '0' + str(ws['B{}'.format(i)].value)
            else:
                pos = str(ws['B{}'.format(i)].value)

            hgpos = hg+pos
            posd_inst.hgpos = int(hgpos)

            if ws['C{}'.format(i)].value:
                posd_inst.description = ws['C{}'.format(i)].value
            else:
                posd_inst.description = ''
            if ws['D{}'.format(i)].value:
                posd_inst.hours = ws['D{}'.format(i)].value
            else:
                posd_inst.hours = ''
            if ws['E{}'.format(i)].value:
                posd_inst.cost = ws['E{}'.format(i)].value
            else:
                posd_inst.cost = ''
            if ws['F{}'.format(i)].value:
                posd_inst.pd_hours = ws['F{}'.format(i)].value
            else:
                posd_inst.pd_hours = None
            if ws['G{}'.format(i)].value:
                posd_inst.pd_cost = ws['G{}'.format(i)].value
            else:
                posd_inst.pd_cost = None

            posd_inst.pk = None
            posd_inst.save()

    except:
        r['code']=1
        r['message']= 'Integrity Error, Project ID already exist in database\n'
        return r

    return r

def upload(request):
    '''
    KBSUMME sheet will be read into database. First check if form request 'POST' otherwise error or
    just refresh (update) browser. Main database entry will happen in function upT3000Func.
    '''

    customer_obj = Customer.objects

    if request.method == 'POST':
        if request.FILES and request.POST.get('customer') and request.POST.get('pid')\
        and request.POST.get('proptype'):
            formdict = {}
            formdict['box_calc'] = request.POST.get('box_calc')
            formdict['box_stckl'] = request.POST.get('box_stckl')
            formdict['customer'] = request.POST.get('customer')
            formdict['pid'] = request.POST.get('pid')
            formdict['proptype'] = request.POST.get('proptype')

            myfile = request.FILES['myfile']
            fs = FileSystemStorage()
            fname = rename(myfile.name)
            filename = fs.save(fname, myfile)
            uploaded_file_url = fs.url(filename)
            r = {}
            r = pduploadFunc(uploaded_file_url, formdict)

            if r['code'] == 1:
                return render(request, 'kbsumme/upload.html', {
                'error':r['message'], 'kbmetaobj':kbmeta_objects, 'customerobj':customer_obj
                } )
            else:
                query_obj = t3000_objects.filter(kbmeta__pid__iexact=str(formdict['pid']))
                return render(request, 'kbsumme/upload.html', {
        'uploaded_file_url': uploaded_file_url,
        'success': 'File successful loaded', 'kbmetaobj':kbmeta_objects, 'customerobj':customer_obj})


        elif (bool(request.FILES.get('myfile_calc', False)) == True) and (bool(int(request.POST['pid_calc'])) == True):
            myfile = request.FILES['myfile_calc']
            pid = request.POST['pid_calc']
            fs = FileSystemStorage()
            fname = rename(myfile.name)
            filename = fs.save(fname, myfile)
            uploaded_file_url = fs.url(filename)
            r = {}
            r = pdupPBBFunc(uploaded_file_url, pid)
            #query_pid = kbmeta_objects.get(pid__iexact=str(pid))
            if r['code'] == 1:
                return render(request, 'kbsumme/upload.html', {'error':r['message'],\
                'kbmetaobj':kbmeta_objects, 'customerobj':customer_obj})
            else:
                return render(request, 'kbsumme/upload.html', {'kbmetaobj':kbmeta_objects, 'customerobj':customer_obj})


        elif (bool(request.FILES.get('myfile_stck', False)) == True) and (bool(int(request.POST['pid_stck'])) == True):
            myfile = request.FILES['myfile_stck']
            pid = request.POST['pid_stck']
            fs = FileSystemStorage()
            fname = rename(myfile.name)
            filename = fs.save(fname, myfile)
            uploaded_file_url = fs.url(filename)
            r = {}
            r = pdupStueckliste(uploaded_file_url, pid)
            if r['code'] == 1:
                return render(request, 'kbsumme/upload.html', {'error':r['message'],\
                'kbmetaobj':kbmeta_objects, 'customerobj':customer_obj})
            else:
                return render(request, 'kbsumme/upload.html', {'kbmetaobj':kbmeta_objects, 'customerobj':customer_obj})


        else:
            return render(request, 'kbsumme/upload.html', {
            'error': 'Please select a file to upload and check the other required fieldelements',\
             'kbmetaobj':kbmeta_objects, 'customerobj':customer_obj})

    else:
        return render(request, 'kbsumme/upload.html', \
        {'kbmetaobj':kbmeta_objects, 'customerobj':customer_obj})

def rename(name):
    name = name.replace(' ', '')
    name = name.replace('&', '')
    name = name.replace(',', '')
    return name

def pduploadFunc(file, formdict):
    '''Same as uploadFunc but with pandas library'''

    kbmeta_inst = KbMeta()
    posd_objects = Posd.objects
    r = {'code':0, 'message':''}

    #try:
    pd_kbsumme = pd.read_excel(BASE_DIR+file, sheet_name='KBSUMME', header=None, na_filter=False)
    #except:
    #    r['code'] = 1
    #    r['message'] = 'KBSUMME could not be loaded'
    #    return r

    if formdict['box_stckl']:
        try:
            pd_stckl = pd.read_excel(BASE_DIR+file, sheet_name='Stueckliste', header=None, na_filter=False)
        except:
            r['code'] = 1
            r['message'] = 'Stueckliste could not be loaded'
            return r

        stueckliste_objects = Stueckliste.objects
        query_stueckliste = stueckliste_objects.filter(kbmeta__pid__iexact=formdict['pid'])

        if query_stueckliste:
            print('Eintrag in Stueckliste vorhanden')
            r['code'] = 1
            r['message'] = 'Stueckliste of this project already in database'
            return r


    if formdict['box_calc']:
        try:
            pd_calc = pd.read_excel(BASE_DIR+file, sheet_name='Calculation', header=None, na_filter=False)
        except:
            r['code'] = 1
            r['message'] = 'Calculation could not be loaded'
            return r

        pbbmeta_objects = PbbMeta.objects
        query_pbbmeta = pbbmeta_objects.filter(kbmeta__pid__iexact=formdict['pid'])

        if query_pbbmeta:
            print('Eintrag in Calculation vorhanden')
            r['code'] = 1
            r['message'] = 'Project calculation sheet already in database'
            return r

    #Fill in the META data to kbmeta database
    #PID, quotno, dateupload, filename cant be empty NULL or not allwed to be emtpy:
    kbmeta_inst.pid = formdict['pid'] #ws['G2'].value
    kbmeta_inst.quotno = pd_kbsumme.iloc[1,6]
    kbmeta_inst.dateupload = timezone.datetime.now()
    kbmeta_inst.filename = file
    kbmeta_inst.phase = formdict['proptype']
    kbmeta_inst.customer = formdict['customer'] #ws['G4'].value
    kbmeta_inst.kbsumme = True
    if formdict['box_stckl']:
        kbmeta_inst.stckliste = True
    else:
        kbmeta_inst.stckliste = False

    if formdict['box_calc']:
        kbmeta_inst.pbb = True
    else:
        kbmeta_inst.pbb = False

    #Values which can be empty. In this case value will output NULL, which is set to not allowed.
    #For char we only accept emtpy string, otherwise double meaning with NULL and empty string will be the case.
    if pd_kbsumme.iloc[4,18]:
        kbmeta_inst.klaversion = str(pd_kbsumme.iloc[4,18])
    if pd_kbsumme.iloc[11,18]:
        kbmeta_inst.calcbase = str(pd_kbsumme.iloc[11,18])
    if pd_kbsumme.iloc[12,18]:
        kbmeta_inst.contractbase = str(pd_kbsumme.iloc[12,18])
    if pd_kbsumme.iloc[2,6]:
        kbmeta_inst.projname = str(pd_kbsumme.iloc[2,6])
    if pd_kbsumme.iloc[4,6]:
        kbmeta_inst.endcustomer = str(pd_kbsumme.iloc[4,6])
    if pd_kbsumme.iloc[3,18]:
        dt = parse(str(pd_kbsumme.iloc[3,18]))
        kbmeta_inst.datecalc = str(dt.strftime('%Y-%m-%d'))

    kbmeta_inst.pk = None
    kbmeta_inst.kbsumme = True
    kbmeta_inst.save()

    for item in posd_objects.all():
        if item.hours:
            t3000_inst.hours = float(pd_kbsumme.iloc[item.pd_hours,10])
        if item.cost:
            t3000_inst.cost = float(pd_kbsumme.iloc[item.pd_cost,12])
        t3000_inst.kbmeta = kbmeta_inst
        t3000_inst.posd = item
        t3000_inst.pk = None
        t3000_inst.save()


    if formdict['box_stckl']:
        r = pdupStueckliste(file, kbmeta_inst.pid)

    if formdict['box_calc']:
        r = pdupPBBFunc(file, kbmeta_inst.pid)

    # This doesnt work. R will be overwritten every time. Dictonaries have
    # to collect-append the messages.

    createDetailProjectGraphic(formdict['pid'])
    return r


def pdupStueckliste(file,pid):
    r = {'code':0, 'message':''}
    stueckliste_objects = Stueckliste.objects
    query_stueckliste = stueckliste_objects.filter(kbmeta__pid__iexact=pid)

    if query_stueckliste:
        print('Eintrag vorhanden')
        r['code'] = 1
        r['message'] = 'Stueckliste for this project already in database'
        return r

    try:
        pd_stckl = pd.read_excel(BASE_DIR+file, sheet_name='Stueckliste')
    except:
        r['code'] = 1
        r['message'] = 'Stueckliste could not be loaded'
        return r

    kbmeta_inst = KbMeta()
    kbmeta_objects = KbMeta.objects
    query_kb = kbmeta_objects.get(pid__iexact=str(pid))
    stueckl_inst = Stueckliste()

    for rw in range(len(pd_stckl)):
        stueckl_inst.module = pd_stckl['Module'].iloc[rw]
        stueckl_inst.sheet = pd_stckl['Sheet'].iloc[rw]
        stueckl_inst.caname = pd_stckl['CA-name'].iloc[rw]
        stueckl_inst.description = pd_stckl['Description'].iloc[rw]
        stueckl_inst.mlfb = pd_stckl['MLFB'].iloc[rw]
        stueckl_inst.qty = pd_stckl['Quantity'].iloc[rw]
        stueckl_inst.single_cost = pd_stckl['Single costs'].iloc[rw]
        stueckl_inst.typical = pd_stckl['Typical'].iloc[rw]
        stueckl_inst.total_cost = pd_stckl['Total costs'].iloc[rw]


        stueckl_inst.kbmeta = query_kb
        stueckl_inst.pk = None
        stueckl_inst.kbmeta.stckliste = True #TODO Does not set the boolean to True in kbmeta. Right code required
        stueckl_inst.save()

    return r


def projectStat(request, pid='205819'):
    obj = get_object_or_404(kbmeta_objects, pid__iexact=str(pid))
    createDetailProjectGraphic(pid)
    return render(request, 'kbsumme/projectStat.html', {'obj':obj})

def projectlist(request):
    return render(request, 'kbsumme/projectlist.html', {'kbmeta_obj':kbmeta_objects})

def price(request):
    query_firm = kbmeta_objects.filter(phase__iexact='')
    pid_list = []
    for query in query_firm:
        pid_list.append(query.pid)

    if request.method == 'POST':
        for query in query_firm:
            if request.POST[str(query.pid)]:
                kbmeta_inst = kbmeta_objects.get(pid__iexact=str(query.pid))
                kbmeta_inst.aprice = float(request.POST[str(query.pid)])
                kbmeta_inst.save()
    return render(request, 'kbsumme/price.html', {'kbmetaobj':query_firm})


def iocount(request):
    iograph()
    return render(request, 'kbsumme/iostat.html')

def upPBB(request):
    if request.method == 'POST':
        if (bool(request.FILES.get('myfile', False)) == True) and (bool(int(request.POST['pid'])) == True):
            myfile = request.FILES['myfile']
            pid = request.POST['pid']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name.replace(" ", ""), myfile)
            uploaded_file_url = fs.url(filename)
            r = {}
            r = upPBBFunc(uploaded_file_url, pid)
            #query_pid = kbmeta_objects.get(pid__iexact=str(pid))
            if r['code'] == 1:
                return render(request, 'kbsumme/upPBB.html', {'error':r['message'],'kbmetaobj':kbmeta_objects})
            else:
                return render(request, 'kbsumme/upPBB.html', {'kbmetaobj':kbmeta_objects})

        else:
            return render(request, 'kbsumme/upPBB.html', {'error':'Please select both\
            PID and File!', 'kbmetaobj':kbmeta_objects})

    else:
        return render(request, 'kbsumme/upPBB.html', {'kbmetaobj':kbmeta_objects})

def pdupPBBFunc(file, pid):
    '''
    Read in the calculation sheet. First right word is searched based on word
    'Business'. Linenumber is sotred in headline. Afterwards the 'Indidvidual Amount'
    field is searched and right column will be stored var col_total.

    Once found, values will be stored to database.
    '''


    r = {'code':0, 'message':''}

    # Check if calculation sheet is not uploaded already
    pbbmeta_objects = PbbMeta.objects
    query_pbbmeta = pbbmeta_objects.filter(kbmeta__pid__iexact=pid)

    if query_pbbmeta:
        print('Eintrag vorhanden')
        r['code'] = 1
        r['message'] = 'Project calculation sheet already in database'
        return r

    # --------------------------------------


    kbmeta_inst = KbMeta()
    kbmeta_objects = KbMeta.objects
    query_kb = kbmeta_objects.get(pid__iexact=str(pid))



    try:
        pd_calc = pd.read_excel(BASE_DIR+file, sheet_name='Calculation', header=None, na_filter=False)
    except:
        r['code'] = 1
        r['message'] = 'Workbook or worksheet could not be loaded'
        return r

    #Search for the right line in the calculation sheet. Once found remember
    #line number in var headline

    pbb_inst = Pbb()
    #Find cell for offerno from (0,0) to (5,end)
    i = 0
    chk_offer = 0
    chk_fx = 0
    for col in range(len(pd_calc.columns)):
        #Goes through each column A to Z and do following:
        for row in range(13):
            #Goes through each row
            if 'Offer-Nr' in str(pd_calc.iloc[row,col]):
                cell_offerno = str(row) + ',' + str(col)
                chk_offer += 1
            if 'Basic rate' in str(pd_calc.iloc[row,col]):
                cell_forex = str(row) + ',' + str(col)
                chk_fx += 1

    # ================= CHECK both pbb and pbbmeta fields before saving to DB ========================
    '''
    if chk_offer == 0:
        r['code'] = 1
        r['message'] = 'Offer-No cell could not be found'
        return r


    e_pid = ws['{}'.format(cell_offerno)].value
    if e_pid == pid:
        r['code'] = 1
        r['message'] = 'Offerno in Calculation and PID do not match'
        return r
    '''

    row = 0
    i=0
    while i < len(pd_calc):
        i+=1
        if 'Business' in str(pd_calc.iloc[i,0]):
            row = i
            break
    #If row with 'Business' could not be found. Print message to user
    if row == 0:
        r['code'] = 1
        r['message'] = 'Could not find headline "Business" in the sheet calculation'
        return r


    #Search for the right column in the calculation sheet. Once found remember db_column
    #in var col_total.
    col = 0
    count = 0
    while count < 27:
        if 'Individual' in str(pd_calc.iloc[row,count]):
            col = count
            break
        count += 1


    # ========================================= END OF CHECK cells ===============================

    pbbmeta_inst = PbbMeta()
    pbbmeta_inst.offer_no = pid #ws['{}'.format(cell_offerno)].value
    pbbmeta_inst.projname = pd_calc.iloc[2,1] #Optimze: Search for field projectname
    pbbmeta_inst.customer = pd_calc.iloc[3,1] #Optimize: Search for field customer

    if chk_fx == 1:
        pbbmeta_inst.fx_rate = pd_calc.iloc[cell_forex]

    #TODO: Read in sheets without below value. Afterwards create html page to let somebody insert manually.

    #pbbmeta_inst.country_inst =
    #pbbmeta_inst.planttype =
    #pbbmeta_inst.qty_units =
    #pbbmeta_inst.total_output =
    #pbbmeta_inst.endcustomer =
    #pbbmeta_inst.bidmanager =

    pbbmeta_inst.total_surcharges = pd_calc.iloc[row+7, col]
    pbbmeta_inst.total_includings = pd_calc.iloc[row+16, col]
    pbbmeta_inst.total_sm = pd_calc.iloc[row+19, col]
    pbbmeta_inst.total_cost = pd_calc.iloc[row+23, col]
    pbbmeta_inst.total_price_euro = pd_calc.iloc[row+26, col]
    pbbmeta_inst.total_price_fx = pd_calc.iloc[row+26, col+1]

    pbbmeta_inst.pk = None
    pbbmeta_inst.kbmeta = query_kb
    pbbmeta_inst.kbmeta.pbb = True
    pbbmeta_inst.save()


        #Steht im Feld Single kost etwas drin und steht im B#headline 'T3000
        #oder Hardware' dann lese die Werte aus Spalte B ein.

    offset = [1,3,4,5,9,10,11,12,13,14,17,18,23,26,31]
    i=0
    while i < (col-2): #Loop all columns
        i+=1
        if pd_calc.iloc[row+offset[13], i]: #Check if field single cost contains some value
            if ('T3000' in pd_calc.iloc[row, i]):
                pbb_inst.kind_of_business = 'T3000 Hardware'
            elif ('Engineering' in pd_calc.iloc[row, i]):
                pbb_inst.kind_of_business = 'SLI Engineering'
            elif ('PM' in pd_calc.iloc[row, i]) or ('CPM' in pd_calc.iloc[row, i]):
                pbb_inst.kind_of_business = 'PM/CPM'
            elif ('Furniture' in pd_calc.iloc[row, i]):
                pbb_inst.kind_of_business = 'Furniture, Printer, Monitor'
            else:
                pbb_inst.kind_of_business = pd_calc.iloc[row, i]

            pbb_inst.entity = pd_calc.iloc[row+offset[0],i]

            if pd_calc.iloc[row+offset[1],i]:
                pbb_inst.escalation = pd_calc.iloc[row+offset[1],i]
            else:
                pbb_inst.escalation = None
            if pd_calc.iloc[row+offset[2],i]:
                pbb_inst.transportation = pd_calc.iloc[row+offset[2],i]
            else:
                pbb_inst.transportation = None
            if pd_calc.iloc[row+offset[3],i]:
                pbb_inst.custom_tax = pd_calc.iloc[row+offset[3],i]
            else:
                pbb_inst.custom_tax = None
            if pd_calc.iloc[row+offset[4],i]:
                pbb_inst.fx_gain_loss = pd_calc.iloc[row+offset[4],i]
            else:
                pbb_inst.fx_gain_loss = None
            if pd_calc.iloc[row+offset[5],i]:
                pbb_inst.financing = pd_calc.iloc[row+offset[5],i]
            else:
                pbb_inst.financing = None
            if pd_calc.iloc[row+offset[6],i]:
                pbb_inst.insurance = pd_calc.iloc[row+offset[6],i]
            else:
                pbb_inst.insurance = None
            if pd_calc.iloc[row+offset[7],i]:
                pbb_inst.bank_charges = pd_calc.iloc[row+offset[7],i]
            else:
                pbb_inst.bank_charges = None
            if pd_calc.iloc[row+offset[8],i]:
                pbb_inst.technical_risk = pd_calc.iloc[row+offset[8],i]
            else:
                pbb_inst.technical_risk = None
            if pd_calc.iloc[row+offset[9],i]:
                pbb_inst.warranty = pd_calc.iloc[row+offset[9],i]
            else:
                pbb_inst.warranty = None
            if pd_calc.iloc[row+offset[10],i]:
                pbb_inst.sales_oh = pd_calc.iloc[row+offset[10],i]
            else:
                pbb_inst.sales_oh = None
            if pd_calc.iloc[row+offset[11],i]:
                pbb_inst.net_margin = pd_calc.iloc[row+offset[11],i]
            else:
                pbb_inst.net_margin = None
            if pd_calc.iloc[row+offset[12],i]:
                pbb_inst.single_cost_euro = pd_calc.iloc[row+offset[12],i]
            else:
                pbb_inst.single_cost_euro = None
            if pd_calc.iloc[row+offset[13],i]:
                pbb_inst.price = pd_calc.iloc[row+offset[13],i]
            else:
                pbb_inst.price = None
            if pd_calc.iloc[row+offset[14],i]:
                pbb_inst.price_nego = pd_calc.iloc[row+offset[14],i]
            else:
                pbb_inst.price_nego = None


            pbb_inst.pbbmeta = pbbmeta_inst
            pbb_inst.pk = None
            pbb_inst.save()

    r['code'] = 0
    r['message'] = ''
    return r
